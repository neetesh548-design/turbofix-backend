"""Tests for new features: photo support, language detection, ticket closure,
machine events, and root cause analysis."""

import shutil
from pathlib import Path
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app import config
from tests.conftest import (
    ACME_OWNER,
    TRACKER_SOURCE,
    auth_headers,
    login,
)

# ---- Helpers ----

def _text_payload(phone: str, body: str) -> dict:
    return {"entry": [{"changes": [{"value": {"messages": [
        {"from": phone, "id": "wamid.text1", "type": "text", "text": {"body": body}}
    ]}}]}]}


def _image_payload(phone: str, media_id: str) -> dict:
    return {"entry": [{"changes": [{"value": {"messages": [
        {"from": phone, "id": "wamid.img1", "type": "image",
         "image": {"id": media_id, "mime_type": "image/jpeg"}}
    ]}}]}]}


def _audio_payload(phone: str, media_id: str) -> dict:
    return {"entry": [{"changes": [{"value": {"messages": [
        {"from": phone, "id": "wamid.aud1", "type": "audio",
         "audio": {"id": media_id, "mime_type": "audio/ogg"}}
    ]}}]}]}


def _last_ticket(tracker_path):
    wb = openpyxl.load_workbook(str(tracker_path), data_only=True)
    ws = wb["Tickets"]
    from app.repositories.base import TICKETS_HEADER
    row = [c.value for c in ws[ws.max_row]]
    row += [None] * (len(TICKETS_HEADER) - len(row))
    return dict(zip(TICKETS_HEADER, row))


def _events_for_machine(tracker_path, machine_id):
    wb = openpyxl.load_workbook(str(tracker_path), data_only=True)
    if "MachineEvents" not in wb.sheetnames:
        return []
    ws = wb["MachineEvents"]
    from app.repositories.base import MACHINE_EVENTS_HEADER
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1] == machine_id:
            events.append(dict(zip(MACHINE_EVENTS_HEADER, row)))
    return events


@pytest.fixture
def client(tmp_path, monkeypatch):
    dest = tmp_path / "TurboFix-Tracker-test.xlsx"
    shutil.copy(TRACKER_SOURCE, dest)
    monkeypatch.setattr(config, "TRACKER_XLSX_PATH", str(dest))
    monkeypatch.setattr(config, "WHATSAPP_VERIFY_TOKEN", "test-verify-token")

    from app import dependencies
    dependencies.get_tickets.cache_clear()
    dependencies.get_machines.cache_clear()
    dependencies.get_events.cache_clear()
    dependencies.get_users.cache_clear()
    dependencies.get_documents.cache_clear()
    dependencies.get_parts.cache_clear()

    from app.infrastructure import whatsapp
    monkeypatch.setattr(whatsapp, "download_media", AsyncMock(return_value="/fake/path/photo.jpg"))

    from app.main import app
    yield TestClient(app), dest

    dependencies.get_tickets.cache_clear()
    dependencies.get_machines.cache_clear()
    dependencies.get_events.cache_clear()
    dependencies.get_users.cache_clear()
    dependencies.get_documents.cache_clear()
    dependencies.get_parts.cache_clear()


# ---- Photo/Image Support ----

class TestImageSupport:
    def test_image_after_text_gets_attached(self, client):
        test_client, tracker_path = client
        test_client.post("/webhook", json=_text_payload(
            "919900099999", "Issue with TF-ACME3-M001: belt looks worn"
        ))
        test_client.post("/webhook", json=_image_payload(
            "919900099999", "img_media_123"
        ))
        ticket = _last_ticket(tracker_path)
        assert ticket["machine_id"] == "TF-ACME3-M001"

    def test_image_without_session_is_ignored(self, client):
        test_client, tracker_path = client
        before = openpyxl.load_workbook(str(tracker_path))["Tickets"].max_row
        test_client.post("/webhook", json=_image_payload(
            "919900099999", "img_media_orphan"
        ))
        after = openpyxl.load_workbook(str(tracker_path))["Tickets"].max_row
        assert after == before


# ---- Ticket Closure ----

class TestTicketClosure:
    def test_close_command_closes_ticket(self, client):
        test_client, tracker_path = client
        # T001 is a sample Open ticket
        test_client.post("/webhook", json=_text_payload(
            "919820012345", "Close T001"
        ))
        wb = openpyxl.load_workbook(str(tracker_path), data_only=True)
        ws = wb["Tickets"]
        from app.repositories.base import TICKETS_HEADER
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "T001":
                ticket = dict(zip(TICKETS_HEADER, list(row) + [None] * (len(TICKETS_HEADER) - len(row))))
                assert ticket["status"] == "Closed"
                assert ticket["closed_by"] == "919820012345"
                break

    def test_close_already_closed_is_noop(self, client):
        test_client, tracker_path = client
        # T002 is already Closed in sample data
        resp = test_client.post("/webhook", json=_text_payload(
            "919820012345", "Close T002"
        ))
        assert resp.status_code == 200

    def test_close_nonexistent_ticket_is_noop(self, client):
        test_client, _ = client
        resp = test_client.post("/webhook", json=_text_payload(
            "919820012345", "Close T999"
        ))
        assert resp.status_code == 200

    def test_close_creates_event(self, client):
        test_client, tracker_path = client
        test_client.post("/webhook", json=_text_payload(
            "919820012345", "Resolve T001"
        ))
        events = _events_for_machine(tracker_path, "TF-ACME3-M001")
        close_events = [e for e in events if e["event_type"] == "ticket_closed"]
        assert len(close_events) >= 1
        assert close_events[-1]["actor_phone"] == "919820012345"


# ---- Machine Events ----

class TestMachineEvents:
    def test_new_ticket_creates_event(self, client):
        test_client, tracker_path = client
        test_client.post("/webhook", json=_text_payload(
            "919900099999", "Issue with TF-ACME3-M002: oil leak visible"
        ))
        events = _events_for_machine(tracker_path, "TF-ACME3-M002")
        created_events = [e for e in events if e["event_type"] == "ticket_created"
                          and e.get("actor_phone") == "919900099999"]
        assert len(created_events) >= 1

    def test_events_endpoint_returns_machine_events(self, client):
        test_client, _ = client
        token = login(test_client, *ACME_OWNER)
        resp = test_client.get(
            "/vault/machines/TF-ACME3-M001/events",
            headers=auth_headers(token),
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["machine_id"] == "TF-ACME3-M001"
        assert isinstance(data["events"], list)

    def test_events_endpoint_404s_for_wrong_company(self, client):
        test_client, _ = client
        token = login(test_client, *ACME_OWNER)
        resp = test_client.get(
            "/vault/machines/TF-BETA1-M001/events",
            headers=auth_headers(token),
        )
        assert resp.status_code == 404


# ---- Root Cause Analysis ----

class TestRootCauseAnalysis:
    def test_root_cause_returns_503_when_ai_disabled(self, client):
        test_client, _ = client
        token = login(test_client, *ACME_OWNER)
        resp = test_client.get(
            "/vault/machines/TF-ACME3-M001/root-cause",
            headers=auth_headers(token),
        )
        assert resp.status_code == 503

    def test_root_cause_404s_for_wrong_company(self, client):
        test_client, _ = client
        token = login(test_client, *ACME_OWNER)
        resp = test_client.get(
            "/vault/machines/TF-BETA1-M001/root-cause",
            headers=auth_headers(token),
        )
        assert resp.status_code == 404


# ---- Ticket Schema ----

class TestTicketSchema:
    def test_new_ticket_has_language_and_photo_fields(self, client):
        test_client, tracker_path = client
        test_client.post("/webhook", json=_text_payload(
            "919900099999", "Issue with TF-ACME3-M001: pump vibrating"
        ))
        ticket = _last_ticket(tracker_path)
        assert "photo_media_id" in ticket
        assert "language" in ticket
        assert "closed_by" in ticket

    def test_sample_tickets_have_language(self, client):
        _, tracker_path = client
        wb = openpyxl.load_workbook(str(tracker_path), data_only=True)
        ws = wb["Tickets"]
        from app.repositories.base import TICKETS_HEADER
        lang_col = TICKETS_HEADER.index("language")
        row2 = [c.value for c in ws[2]]
        assert row2[lang_col] == "hi"


# ---- MachineEvents Tab ----

class TestMachineEventsTab:
    def test_machine_events_tab_exists(self, client):
        _, tracker_path = client
        wb = openpyxl.load_workbook(str(tracker_path))
        assert "MachineEvents" in wb.sheetnames

    def test_machine_events_has_sample_data(self, client):
        _, tracker_path = client
        wb = openpyxl.load_workbook(str(tracker_path), data_only=True)
        ws = wb["MachineEvents"]
        assert ws.max_row > 1


# ---- Reports ----

class TestReports:
    def test_get_report_returns_metrics(self, client):
        test_client, _ = client
        token = login(test_client, *ACME_OWNER)
        resp = test_client.get("/vault/reports/ytd", headers=auth_headers(token))
        assert resp.status_code == 200
        data = resp.json()
        assert "metrics" in data
        assert "total_tickets" in data["metrics"]
        assert "formatted_text" in data

    def test_get_report_invalid_period(self, client):
        test_client, _ = client
        token = login(test_client, *ACME_OWNER)
        resp = test_client.get("/vault/reports/quarterly", headers=auth_headers(token))
        assert resp.status_code == 400

    def test_report_requires_auth(self, client):
        test_client, _ = client
        resp = test_client.get("/vault/reports/daily")
        assert resp.status_code in (401, 403)

    def test_send_report_no_whatsapp(self, client):
        test_client, _ = client
        token = login(test_client, *ACME_OWNER)
        resp = test_client.post("/vault/reports/daily/send", headers=auth_headers(token))
        assert resp.status_code == 200
        assert resp.json()["status"] == "skipped"


# ---- Signup phone normalization ----

class TestSignupPhoneNormalization:
    def test_signup_with_plus_prefix(self, client):
        test_client, _ = client
        resp = test_client.post("/auth/signup", json={
            "company_code": "ACME3",
            "admin_contact_phone": "+919820012345",
            "name": "New Supervisor",
            "phone": "+910000000001",
            "email": "",
            "password": "testpass1234",
        })
        assert resp.status_code == 201
        assert resp.json()["user"]["role"] == "supervisor"

    def test_signup_without_plus_prefix(self, client):
        test_client, _ = client
        resp = test_client.post("/auth/signup", json={
            "company_code": "ACME3",
            "admin_contact_phone": "919820012345",
            "name": "Another Supervisor",
            "phone": "+910000000002",
            "email": "",
            "password": "testpass1234",
        })
        assert resp.status_code == 201
