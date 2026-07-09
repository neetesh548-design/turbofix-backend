"""Tests for the local (openpyxl) ticket and machine repositories.

Replaces test_store_local.py — same behaviour, updated to use the SOLID
repository classes instead of the flat store_local module.
"""

import shutil
from pathlib import Path

import openpyxl
import pytest

from app import config
from app.repositories.local.ticket_repo import LocalMachineRepository, LocalTicketRepository

TRACKER_SOURCE = Path(__file__).resolve().parent.parent.parent / "TurboFix-Tracker.xlsx"


@pytest.fixture
def tracker_path(tmp_path):
    dest = tmp_path / "TurboFix-Tracker-test.xlsx"
    shutil.copy(TRACKER_SOURCE, dest)
    return dest


@pytest.fixture
def ticket_repo(tracker_path):
    return LocalTicketRepository(str(tracker_path))


@pytest.fixture
def machine_repo(tracker_path):
    return LocalMachineRepository(str(tracker_path))


def test_load_machines_reads_expected_rows(machine_repo):
    machines = machine_repo.load()
    assert "TF-ACME3-M001" in machines
    assert machines["TF-ACME3-M001"]["company_code"] == "ACME3"
    assert machines["TF-ACME3-M001"]["assigned_technician_phone"] == "+919812340001"
    assert machines["TF-ACME3-M001"]["informed_phones"] == ["+919812340010", "+919812340011"]


def test_load_machines_is_cached_within_ttl(tracker_path):
    repo = LocalMachineRepository(str(tracker_path), cache_ttl=9999)
    first = repo.load()
    # Delete the file — if caching works, the second call still returns data
    tracker_path.unlink()
    second = repo.load()
    assert second == first


def test_load_machines_refreshes_after_ttl_expires(tracker_path):
    repo = LocalMachineRepository(str(tracker_path), cache_ttl=0)
    first = repo.load()
    first["TF-ACME3-M001"]["machine_name"] = "mutated in returned dict"
    second = repo.load()
    assert second["TF-ACME3-M001"]["machine_name"] == "CNC Lathe 1"


def test_append_ticket_adds_a_row(ticket_repo, tracker_path):
    row = {
        "ticket_id": "TTEST001",
        "machine_id": "TF-ACME3-M001",
        "company_code": "ACME3",
        "machine_name": "CNC Lathe 1",
        "reported_at": "2026-07-07 12:00",
        "reporter_phone": "+919900099999",
        "description": "test description",
        "ai_summary": "",
        "urgency": "",
        "status": "Open",
        "closed_at": "",
        "hours_to_fix": "",
        "voice_note_media_id": "",
    }
    ticket_repo.append(row)

    wb = openpyxl.load_workbook(str(tracker_path))
    ws = wb["Tickets"]
    last_row = [c.value for c in ws[ws.max_row]]
    assert last_row[0] == "TTEST001"
    assert last_row[1] == "TF-ACME3-M001"
    assert last_row[6] == "test description"


def test_attach_voice_note_updates_matching_row(ticket_repo, tracker_path):
    ticket_repo.append({
        "ticket_id": "TTEST002", "machine_id": "TF-ACME3-M001", "company_code": "ACME3",
        "machine_name": "CNC Lathe 1", "reported_at": "2026-07-07 12:00",
        "reporter_phone": "+919900099999", "description": "d", "ai_summary": "",
        "urgency": "", "status": "Open", "closed_at": "", "hours_to_fix": "",
        "voice_note_media_id": "",
    })
    found = ticket_repo.attach_voice_note("TTEST002", "wamid.HBgL...")
    assert found is True

    wb = openpyxl.load_workbook(str(tracker_path))
    ws = wb["Tickets"]
    last_row = [c.value for c in ws[ws.max_row]]
    assert last_row[12] == "wamid.HBgL..."


def test_attach_voice_note_returns_false_for_unknown_ticket(ticket_repo):
    assert ticket_repo.attach_voice_note("NOPE", "media123") is False


def test_next_ticket_id_is_unique(ticket_repo):
    a = ticket_repo.next_ticket_id()
    b = ticket_repo.next_ticket_id()
    assert a != b
    assert a.startswith("T")


def test_get_ticket_returns_matching_row(ticket_repo):
    ticket_repo.append({
        "ticket_id": "TTEST003", "machine_id": "TF-ACME3-M001", "company_code": "ACME3",
        "machine_name": "CNC Lathe 1", "reported_at": "2026-07-07 12:00",
        "reporter_phone": "+919900099999", "description": "original description",
        "ai_summary": "", "urgency": "", "status": "Open", "closed_at": "",
        "hours_to_fix": "", "voice_note_media_id": "",
    })
    ticket = ticket_repo.get("TTEST003")
    assert ticket is not None
    assert ticket["description"] == "original description"
    assert ticket["machine_id"] == "TF-ACME3-M001"


def test_get_ticket_returns_none_for_unknown_id(ticket_repo):
    assert ticket_repo.get("NOPE") is None


def test_update_ai_fields_sets_summary_urgency_and_description(ticket_repo):
    ticket_repo.append({
        "ticket_id": "TTEST004", "machine_id": "TF-ACME3-M001", "company_code": "ACME3",
        "machine_name": "CNC Lathe 1", "reported_at": "2026-07-07 12:00",
        "reporter_phone": "+919900099999", "description": "(no description provided)",
        "ai_summary": "", "urgency": "", "status": "Open", "closed_at": "",
        "hours_to_fix": "", "voice_note_media_id": "",
    })
    found = ticket_repo.update_ai_fields(
        "TTEST004",
        ai_summary="Likely cause: worn bearing | Suggested action: inspect bearing",
        urgency="High",
        description="spindle making loud noise",
    )
    assert found is True

    ticket = ticket_repo.get("TTEST004")
    assert ticket["ai_summary"] == "Likely cause: worn bearing | Suggested action: inspect bearing"
    assert ticket["urgency"] == "High"
    assert ticket["description"] == "spindle making loud noise"


def test_update_ai_fields_returns_false_for_unknown_ticket(ticket_repo):
    assert ticket_repo.update_ai_fields("NOPE", "summary", "Low") is False
