"""Webhook endpoint tests — updated for the SOLID architecture.

Uses FastAPI's dependency_overrides (the SOLID way) instead of monkeypatching
global module attributes on main.py.
"""

import asyncio
import shutil
from pathlib import Path
from unittest.mock import AsyncMock

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app import config

TRACKER_SOURCE = Path(__file__).resolve().parent.parent.parent / "TurboFix-Tracker.xlsx"


def _text_payload(phone: str, body: str) -> dict:
    return {
        "entry": [{"changes": [{"value": {"messages": [
            {"from": phone, "id": "wamid.text1", "type": "text", "text": {"body": body}}
        ]}}]}]
    }


def _audio_payload(phone: str, media_id: str) -> dict:
    return {
        "entry": [{"changes": [{"value": {"messages": [
            {"from": phone, "id": "wamid.audio1", "type": "audio",
             "audio": {"id": media_id, "mime_type": "audio/ogg"}}
        ]}}]}]
    }


@pytest.fixture
def client(tmp_path, monkeypatch):
    dest = tmp_path / "TurboFix-Tracker-test.xlsx"
    shutil.copy(TRACKER_SOURCE, dest)
    monkeypatch.setattr(config, "TRACKER_XLSX_PATH", str(dest))
    monkeypatch.setattr(config, "WHATSAPP_VERIFY_TOKEN", "test-verify-token")

    # Clear DI caches so they pick up the monkeypatched config.
    from app import dependencies
    dependencies.get_tickets.cache_clear()
    dependencies.get_machines.cache_clear()
    dependencies.get_users.cache_clear()
    dependencies.get_documents.cache_clear()
    dependencies.get_parts.cache_clear()

    # Patch the whatsapp module in infrastructure so download_media doesn't make HTTP calls.
    from app.infrastructure import whatsapp
    monkeypatch.setattr(whatsapp, "download_media", AsyncMock(return_value="/fake/path/audio.ogg"))

    from app.main import app

    yield TestClient(app), dest

    # Cleanup
    dependencies.get_tickets.cache_clear()
    dependencies.get_machines.cache_clear()
    dependencies.get_users.cache_clear()
    dependencies.get_documents.cache_clear()
    dependencies.get_parts.cache_clear()


def _last_ticket_row(tracker_path):
    wb = openpyxl.load_workbook(str(tracker_path))
    ws = wb["Tickets"]
    return [c.value for c in ws[ws.max_row]]


# ---------------------------------------------------------------------------
# Webhook verification
# ---------------------------------------------------------------------------

def test_webhook_verification_handshake(client):
    test_client, _ = client
    resp = test_client.get("/webhook", params={
        "hub.mode": "subscribe",
        "hub.verify_token": "test-verify-token",
        "hub.challenge": "12345",
    })
    assert resp.status_code == 200
    assert resp.text == "12345"


def test_webhook_verification_rejects_wrong_token(client):
    test_client, _ = client
    resp = test_client.get("/webhook", params={
        "hub.mode": "subscribe",
        "hub.verify_token": "wrong",
        "hub.challenge": "12345",
    })
    assert resp.status_code == 403


# ---------------------------------------------------------------------------
# Text messages
# ---------------------------------------------------------------------------

def test_text_message_with_known_machine_logs_ticket(client):
    test_client, tracker_path = client
    resp = test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    assert resp.status_code == 200

    row = _last_ticket_row(tracker_path)
    assert row[1] == "TF-ACME3-M001"
    assert row[2] == "ACME3"
    assert row[3] == "CNC Lathe 1"
    assert row[5] == "919900012345"
    assert row[6] == "spindle making loud noise"
    assert row[9] == "Open"
    assert row[12] in ("", None)


def test_text_message_with_unknown_machine_is_ignored(client):
    test_client, tracker_path = client
    before_rows = openpyxl.load_workbook(str(tracker_path))["Tickets"].max_row

    resp = test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ZZZZ-M999: something"
    ))
    assert resp.status_code == 200

    after_rows = openpyxl.load_workbook(str(tracker_path))["Tickets"].max_row
    assert after_rows == before_rows


# ---------------------------------------------------------------------------
# Audio messages
# ---------------------------------------------------------------------------

def test_voice_note_after_text_gets_attached_to_ticket(client):
    test_client, tracker_path = client
    test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    ticket_id_before = _last_ticket_row(tracker_path)[0]

    resp = test_client.post("/webhook", json=_audio_payload("919900012345", "wamid.AUDIO123"))
    assert resp.status_code == 200

    row = _last_ticket_row(tracker_path)
    assert row[0] == ticket_id_before
    assert row[12] == "wamid.AUDIO123"


def test_voice_note_without_prior_text_is_dropped(client):
    test_client, tracker_path = client
    before_rows = openpyxl.load_workbook(str(tracker_path))["Tickets"].max_row

    resp = test_client.post("/webhook", json=_audio_payload("919900099999", "wamid.ORPHAN"))
    assert resp.status_code == 200

    after_rows = openpyxl.load_workbook(str(tracker_path))["Tickets"].max_row
    assert after_rows == before_rows


# ---------------------------------------------------------------------------
# AI summarization (via service-level monkeypatching)
# ---------------------------------------------------------------------------

class _FakeBrief:
    def __init__(self, likely_cause, urgency, suggested_action):
        self.likely_cause = likely_cause
        self.urgency = urgency
        self.suggested_action = suggested_action

    def as_ai_summary(self):
        return f"Likely cause: {self.likely_cause} | Suggested action: {self.suggested_action}"


def test_text_message_with_description_gets_ai_summary(client, monkeypatch):
    test_client, tracker_path = client
    monkeypatch.setattr(config, "OPENAI_API_KEY", "fake-key")

    from app.services import ai_service

    async def fake_summarize_issue(description):
        assert description == "spindle making loud noise"
        return _FakeBrief("worn bearing", "High", "inspect and replace bearing")

    monkeypatch.setattr(ai_service, "summarize_issue", fake_summarize_issue)
    monkeypatch.setattr(ai_service, "ai_enabled", lambda: True)

    resp = test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    assert resp.status_code == 200

    row = _last_ticket_row(tracker_path)
    assert row[7] == "Likely cause: worn bearing | Suggested action: inspect and replace bearing"
    assert row[8] == "High"
    assert row[6] == "spindle making loud noise"


def test_ai_failure_is_swallowed_and_ticket_stays_logged(client, monkeypatch):
    test_client, tracker_path = client

    from app.services import ai_service

    async def failing_summarize(description):
        raise RuntimeError("simulated API outage")

    monkeypatch.setattr(ai_service, "summarize_issue", failing_summarize)
    monkeypatch.setattr(ai_service, "ai_enabled", lambda: True)

    resp = test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    assert resp.status_code == 200

    row = _last_ticket_row(tracker_path)
    assert row[1] == "TF-ACME3-M001"
    assert row[7] in ("", None)  # ai_summary left blank, not crashed


def test_no_ai_keys_skips_ai_but_still_logs_ticket(client, monkeypatch):
    test_client, tracker_path = client
    monkeypatch.setattr(config, "OPENAI_API_KEY", "")
    monkeypatch.setattr(config, "GEMINI_API_KEY", "")

    resp = test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    assert resp.status_code == 200
    row = _last_ticket_row(tracker_path)
    assert row[1] == "TF-ACME3-M001"


# ---------------------------------------------------------------------------
# Fan-out (via fanout_service monkeypatching)
# ---------------------------------------------------------------------------

def _enable_fanout_credentials(monkeypatch):
    monkeypatch.setattr(config, "WHATSAPP_ACCESS_TOKEN", "fake-token")
    monkeypatch.setattr(config, "WHATSAPP_PHONE_NUMBER_ID", "1234567890")


def test_fanout_skipped_without_whatsapp_send_credentials(client, monkeypatch):
    test_client, _ = client
    monkeypatch.setattr(config, "WHATSAPP_ACCESS_TOKEN", "")
    monkeypatch.setattr(config, "WHATSAPP_PHONE_NUMBER_ID", "")

    from app.infrastructure import whatsapp

    called = False

    async def fake_send(to, params):
        nonlocal called
        called = True

    monkeypatch.setattr(whatsapp, "send_template_message", fake_send)

    resp = test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    assert resp.status_code == 200
    assert called is False


def test_text_with_description_triggers_fanout(client, monkeypatch):
    test_client, _ = client
    _enable_fanout_credentials(monkeypatch)

    from app.services import fanout_service

    calls = []

    async def fake_notify(machine, ticket):
        calls.append((machine, ticket))

    monkeypatch.setattr(fanout_service, "notify_ticket", fake_notify)

    resp = test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    assert resp.status_code == 200
    assert len(calls) == 1
    machine, ticket = calls[0]
    assert machine["assigned_technician_phone"] == "+919812340001"
    assert machine["informed_phones"] == ["+919812340010", "+919812340011"]
    assert ticket["machine_id"] == "TF-ACME3-M001"


def test_bare_id_text_does_not_fanout_until_voice_note_arrives(client, monkeypatch):
    test_client, _ = client
    _enable_fanout_credentials(monkeypatch)

    from app.services import fanout_service

    calls = []

    async def fake_notify(machine, ticket):
        calls.append((machine, ticket))

    monkeypatch.setattr(fanout_service, "notify_ticket", fake_notify)

    test_client.post("/webhook", json=_text_payload("919900012345", "TF-ACME3-M001"))
    assert calls == []

    resp = test_client.post("/webhook", json=_audio_payload("919900012345", "wamid.AUDIO1"))
    assert resp.status_code == 200
    assert len(calls) == 1


def test_text_then_voice_note_only_fans_out_once(client, monkeypatch):
    test_client, _ = client
    _enable_fanout_credentials(monkeypatch)

    from app.services import fanout_service

    calls = []

    async def fake_notify(machine, ticket):
        calls.append((machine, ticket))

    monkeypatch.setattr(fanout_service, "notify_ticket", fake_notify)

    test_client.post("/webhook", json=_text_payload(
        "919900012345", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    assert len(calls) == 1

    resp = test_client.post("/webhook", json=_audio_payload("919900012345", "wamid.AUDIO2"))
    assert resp.status_code == 200
    assert len(calls) == 1  # still just the one fan-out from the text message


def test_fanout_fires_even_when_transcription_fails(client, monkeypatch):
    test_client, _ = client
    _enable_fanout_credentials(monkeypatch)

    from app.services import ai_service, fanout_service

    async def failing_transcribe(path):
        raise RuntimeError("simulated transcription outage")

    calls = []

    async def fake_notify(machine, ticket):
        calls.append((machine, ticket))

    monkeypatch.setattr(ai_service, "transcribe_audio", failing_transcribe)
    monkeypatch.setattr(ai_service, "ai_enabled", lambda: True)
    monkeypatch.setattr(fanout_service, "notify_ticket", fake_notify)

    test_client.post("/webhook", json=_text_payload("919900012345", "TF-ACME3-M001"))
    resp = test_client.post("/webhook", json=_audio_payload("919900012345", "wamid.AUDIO3"))

    assert resp.status_code == 200
    assert len(calls) == 1


# ---------------------------------------------------------------------------
# Session sweep
# ---------------------------------------------------------------------------

def test_sweep_fans_out_orphaned_bare_id_session(client, monkeypatch):
    test_client, _ = client
    _enable_fanout_credentials(monkeypatch)

    from app.services import fanout_service
    from app.routers.webhook_router import get_sessions

    calls = []

    async def fake_notify(machine, ticket):
        calls.append((machine, ticket))

    monkeypatch.setattr(fanout_service, "notify_ticket", fake_notify)

    # bare ID, no voice note ever follows
    test_client.post("/webhook", json=_text_payload("919900054321", "TF-ACME3-M001"))
    assert calls == []

    # simulate the session having expired without a follow-up voice note
    sessions = get_sessions()
    session = sessions._sessions["919900054321"]
    session.created_at -= config.SESSION_TTL_SECONDS + 1

    from app import dependencies
    from app.services.ticket_service import sweep_expired_unnotified
    asyncio.run(sweep_expired_unnotified(sessions, dependencies.get_tickets(), dependencies.get_machines()))

    assert len(calls) == 1
    machine, ticket = calls[0]
    assert ticket["machine_id"] == "TF-ACME3-M001"
    assert "919900054321" not in sessions._sessions


def test_sweep_does_not_refanout_already_notified_sessions(client, monkeypatch):
    test_client, _ = client
    _enable_fanout_credentials(monkeypatch)

    from app.services import fanout_service
    from app.routers.webhook_router import get_sessions

    calls = []

    async def fake_notify(machine, ticket):
        calls.append((machine, ticket))

    monkeypatch.setattr(fanout_service, "notify_ticket", fake_notify)

    test_client.post("/webhook", json=_text_payload(
        "919900054322", "Issue with TF-ACME3-M001: spindle making loud noise"
    ))
    assert len(calls) == 1  # already fanned out via typed description

    sessions = get_sessions()
    session = sessions._sessions["919900054322"]
    session.created_at -= config.SESSION_TTL_SECONDS + 1

    from app import dependencies
    from app.services.ticket_service import sweep_expired_unnotified
    asyncio.run(sweep_expired_unnotified(sessions, dependencies.get_tickets(), dependencies.get_machines()))

    assert len(calls) == 1  # sweep did not fire a second, duplicate notification
