import re
import secrets
import threading
import time
from datetime import datetime, timezone
from typing import Dict, Optional

import openpyxl

from app import config

_lock = threading.Lock()

# Harden phase: Machines rarely change compared to message volume, so cache the last
# read rather than re-parsing the whole tab on every incoming message. Keyed by
# tracker path so tests using different tmp_path trackers never see each other's data.
_machines_cache: Optional[Dict[str, dict]] = None
_machines_cache_key: Optional[str] = None
_machines_cache_at: float = 0.0

_MACHINES_HEADER = ["machine_id", "company_code", "machine_name", "location",
                     "assigned_technician_phone", "informed_phone_1", "informed_phone_2",
                     "informed_phone_3", "has_open_tickets"]
_TICKETS_HEADER = ["ticket_id", "machine_id", "company_code", "machine_name", "reported_at",
                    "reporter_phone", "description", "ai_summary", "urgency", "status",
                    "closed_at", "hours_to_fix", "voice_note_media_id"]


class MachineNotFoundError(Exception):
    pass


def load_machines() -> Dict[str, dict]:
    """Returns {machine_id: {company_code, machine_name, assigned_technician_phone, informed_phones}}.
    Cached in-process for MACHINES_CACHE_TTL_SECONDS since Machines changes rarely
    compared to message volume - a change to the tab can take up to that long to
    take effect."""
    global _machines_cache, _machines_cache_key, _machines_cache_at

    now = time.time()
    if (
        _machines_cache is not None
        and _machines_cache_key == config.TRACKER_XLSX_PATH
        and now - _machines_cache_at < config.MACHINES_CACHE_TTL_SECONDS
    ):
        return _machines_cache

    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    ws = wb["Machines"]
    machines = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        machine_id = str(row[0]).strip().upper()
        machines[machine_id] = {
            "company_code": row[1],
            "machine_name": row[2],
            "location": row[3],
            "assigned_technician_phone": row[4],
            "informed_phones": [p for p in (row[5], row[6], row[7]) if p],
        }

    _machines_cache = machines
    _machines_cache_key = config.TRACKER_XLSX_PATH
    _machines_cache_at = now
    return machines


def get_machine(machine_id: str) -> Optional[dict]:
    return load_machines().get(machine_id.upper())


def invalidate_machines_cache() -> None:
    """Forces the next load_machines() call to re-read the tracker instead of
    returning a cached copy - used right after create_machine() so a machine
    added via the Vault UI is immediately visible to the webhook, without
    waiting out MACHINES_CACHE_TTL_SECONDS."""
    global _machines_cache, _machines_cache_key, _machines_cache_at
    _machines_cache = None
    _machines_cache_key = None
    _machines_cache_at = 0.0


def next_machine_code(company_code: str) -> str:
    """Returns the next Mnnn code for a company, e.g. "M003" if M001/M002 already
    exist. Forces a fresh read (not the cache) since this is called rarely - on
    machine creation, not the message-handling hot path - and correctness here
    matters more than avoiding a re-read."""
    invalidate_machines_cache()
    pattern = re.compile(rf"^TF-{re.escape(company_code)}-M(\d+)$", re.IGNORECASE)
    max_n = 0
    for machine_id in load_machines():
        match = pattern.match(machine_id)
        if match:
            max_n = max(max_n, int(match.group(1)))
    return f"M{max_n + 1:03d}"


def create_machine(row: dict) -> None:
    """row keys: machine_id, company_code, machine_name, location,
    assigned_technician_phone, informed_phone_1, informed_phone_2, informed_phone_3.
    has_open_tickets is a formula column (see build_tracker.py) - set here the same
    way so "Machines Down" keeps working for machines added outside build_tracker.py."""
    data_cols = _MACHINES_HEADER[:-1]
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        ws = wb["Machines"]
        ws.append([row.get(col, "") for col in data_cols])
        row_num = ws.max_row
        ws.cell(
            row=row_num, column=len(_MACHINES_HEADER),
            value=f'=COUNTIFS(Tickets!$B:$B,A{row_num},Tickets!$J:$J,"Open")',
        )
        wb.save(config.TRACKER_XLSX_PATH)
    invalidate_machines_cache()


def append_ticket(row: dict) -> None:
    """row keys: ticket_id, machine_id, company_code, machine_name, reported_at,
    reporter_phone, description, ai_summary, urgency, status, closed_at, hours_to_fix,
    voice_note_media_id"""
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        ws = wb["Tickets"]
        ws.append([row.get(col, "") for col in _TICKETS_HEADER])
        wb.save(config.TRACKER_XLSX_PATH)


def attach_voice_note(ticket_id: str, media_id: str) -> bool:
    """Finds the ticket row by ticket_id and sets its voice_note_media_id column.
    Returns True if the ticket was found and updated."""
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        ws = wb["Tickets"]
        media_col = _TICKETS_HEADER.index("voice_note_media_id") + 1
        for row_cells in ws.iter_rows(min_row=2):
            if row_cells[0].value == ticket_id:
                row_cells[media_col - 1].value = media_id
                wb.save(config.TRACKER_XLSX_PATH)
                return True
        return False


def get_ticket(ticket_id: str) -> Optional[dict]:
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    ws = wb["Tickets"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == ticket_id:
            return dict(zip(_TICKETS_HEADER, row))
    return None


def update_ai_fields(ticket_id: str, ai_summary: str, urgency: str, description: Optional[str] = None) -> bool:
    """Sets ai_summary/urgency (and optionally overwrites description, e.g. once a
    voice note has been transcribed) on the matching ticket row. Returns True if found."""
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        ws = wb["Tickets"]
        for row_cells in ws.iter_rows(min_row=2):
            if row_cells[0].value == ticket_id:
                row_cells[_TICKETS_HEADER.index("ai_summary")].value = ai_summary
                row_cells[_TICKETS_HEADER.index("urgency")].value = urgency
                if description is not None:
                    row_cells[_TICKETS_HEADER.index("description")].value = description
                wb.save(config.TRACKER_XLSX_PATH)
                return True
        return False


def next_ticket_id() -> str:
    return f"T{datetime.now(timezone.utc):%Y%m%d%H%M%S}-{secrets.token_hex(2)}"
