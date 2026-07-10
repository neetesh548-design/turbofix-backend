"""Local (openpyxl / Excel) implementations of TicketRepository and MachineRepository.

These are the dev/test defaults — no external credentials needed.
In production TICKET_STORE=sheets selects the Sheets implementations instead.
"""

import re
import threading
import time
from typing import Dict, List, Optional

import openpyxl

from app import config
from app.repositories.base import (
    MACHINE_EVENTS_HEADER,
    MACHINES_HEADER,
    TICKETS_HEADER,
    EventRepository,
    MachineRepository,
    TicketRepository,
    new_event_id,
    new_ticket_id,
)


class LocalTicketRepository(TicketRepository):
    """Reads/writes tickets in the Tickets tab of the local tracker workbook."""

    def __init__(self, xlsx_path: str):
        self._path = xlsx_path
        self._lock = threading.Lock()

    def next_ticket_id(self) -> str:
        return new_ticket_id()

    def append(self, row: dict) -> None:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = wb["Tickets"]
            ws.append([row.get(col, "") for col in TICKETS_HEADER])
            wb.save(self._path)

    def get(self, ticket_id: str) -> Optional[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        ws = wb["Tickets"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == ticket_id:
                return dict(zip(TICKETS_HEADER, row))
        return None

    def attach_voice_note(self, ticket_id: str, media_id: str) -> bool:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = wb["Tickets"]
            media_col = TICKETS_HEADER.index("voice_note_media_id") + 1
            for row_cells in ws.iter_rows(min_row=2):
                if row_cells[0].value == ticket_id:
                    row_cells[media_col - 1].value = media_id
                    wb.save(self._path)
                    return True
        return False

    def update_ai_fields(
        self,
        ticket_id: str,
        ai_summary: str,
        urgency: str,
        description: Optional[str] = None,
    ) -> bool:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = wb["Tickets"]
            for row_cells in ws.iter_rows(min_row=2):
                if row_cells[0].value == ticket_id:
                    row_cells[TICKETS_HEADER.index("ai_summary")].value = ai_summary
                    row_cells[TICKETS_HEADER.index("urgency")].value = urgency
                    if description is not None:
                        row_cells[TICKETS_HEADER.index("description")].value = description
                    wb.save(self._path)
                    return True
        return False

    def get_company_tickets(self, company_code: str) -> List[dict]:
        with self._lock:
            wb = openpyxl.load_workbook(self._path, data_only=True)
            ws = wb["Tickets"]
            tickets = []
            for row_cells in ws.iter_rows(min_row=2, values_only=False):
                company = row_cells[TICKETS_HEADER.index("company_code")].value
                if company == company_code:
                    tickets.append(
                        {col: row_cells[i].value for i, col in enumerate(TICKETS_HEADER)}
                    )
            return tickets

    def attach_photo(self, ticket_id: str, media_id: str) -> bool:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = wb["Tickets"]
            col_idx = TICKETS_HEADER.index("photo_media_id")
            for row_cells in ws.iter_rows(min_row=2):
                if row_cells[0].value == ticket_id:
                    row_cells[col_idx].value = media_id
                    wb.save(self._path)
                    return True
        return False

    def update_language(self, ticket_id: str, language: str) -> bool:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = wb["Tickets"]
            col_idx = TICKETS_HEADER.index("language")
            for row_cells in ws.iter_rows(min_row=2):
                if row_cells[0].value == ticket_id:
                    row_cells[col_idx].value = language
                    wb.save(self._path)
                    return True
        return False

    def close_ticket(self, ticket_id: str, closed_by: str) -> bool:
        from datetime import datetime, timezone
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = wb["Tickets"]
            for row_cells in ws.iter_rows(min_row=2):
                if row_cells[0].value == ticket_id:
                    row_cells[TICKETS_HEADER.index("status")].value = "Closed"
                    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
                    row_cells[TICKETS_HEADER.index("closed_at")].value = now
                    row_cells[TICKETS_HEADER.index("closed_by")].value = closed_by
                    wb.save(self._path)
                    return True
        return False

    def find_by_id_prefix(self, prefix: str) -> Optional[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        ws = wb["Tickets"]
        prefix_upper = prefix.upper()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and str(row[0]).upper().startswith(prefix_upper):
                return dict(zip(TICKETS_HEADER, row))
        return None


class LocalEventRepository(EventRepository):
    """Reads/writes events in the MachineEvents tab of the local tracker workbook."""

    def __init__(self, xlsx_path: str):
        self._path = xlsx_path
        self._lock = threading.Lock()

    def _ensure_tab(self, wb):
        if "MachineEvents" not in wb.sheetnames:
            ws = wb.create_sheet("MachineEvents")
            ws.append(MACHINE_EVENTS_HEADER)
            return ws
        return wb["MachineEvents"]

    def append(self, row: dict) -> None:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = self._ensure_tab(wb)
            ws.append([row.get(col, "") for col in MACHINE_EVENTS_HEADER])
            wb.save(self._path)

    def get_machine_events(self, machine_id: str) -> List[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if "MachineEvents" not in wb.sheetnames:
            return []
        ws = wb["MachineEvents"]
        events = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1] == machine_id:
                events.append(dict(zip(MACHINE_EVENTS_HEADER, row)))
        return events

    def get_company_events(self, company_code: str) -> List[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if "MachineEvents" not in wb.sheetnames:
            return []
        ws = wb["MachineEvents"]
        events = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[2] == company_code:
                events.append(dict(zip(MACHINE_EVENTS_HEADER, row)))
        return events


class LocalMachineRepository(MachineRepository):
    """Reads/writes machines in the Machines tab of the local tracker workbook.

    Maintains an in-process cache keyed by xlsx path, same as the legacy
    store_local.py, so machine lookups on the message-handling hot path
    don't re-parse the workbook every time.
    """

    def __init__(self, xlsx_path: str, cache_ttl: int = 60):
        self._path = xlsx_path
        self._cache_ttl = cache_ttl
        self._lock = threading.Lock()
        self._cache: Optional[Dict[str, dict]] = None
        self._cache_at: float = 0.0

    def load(self) -> Dict[str, dict]:
        now = time.time()
        if self._cache is not None and now - self._cache_at < self._cache_ttl:
            return self._cache

        wb = openpyxl.load_workbook(self._path, data_only=True)
        ws = wb["Machines"]
        machines: Dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            machine_id = str(row[0]).strip().upper()
            machines[machine_id] = {
                "company_code": row[1],
                "machine_name": row[2],
                "location": row[3],
                "assigned_technician_phone": row[4],
                "informed_phones": [p for p in (row[5], row[6], row[7]) if p],
                "supervisor_id": row[8] if len(row) > 8 and row[8] is not None else "",
                "last_activity_at": str(row[10]) if len(row) > 10 and row[10] is not None else "",
            }

        self._cache = machines
        self._cache_at = now
        return machines

    def get(self, machine_id: str) -> Optional[dict]:
        return self.load().get(machine_id.upper())

    def create(self, row: dict) -> None:
        data_cols = MACHINES_HEADER[:-1]  # exclude has_open_tickets (formula column)
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = wb["Machines"]
            ws.append([row.get(col, "") for col in data_cols])
            row_num = ws.max_row
            ws.cell(
                row=row_num, column=len(MACHINES_HEADER),
                value=f'=COUNTIFS(Tickets!$B:$B,A{row_num},Tickets!$J:$J,"Open")',
            )
            wb.save(self._path)
        self.invalidate_cache()

    def invalidate_cache(self) -> None:
        self._cache = None
        self._cache_at = 0.0

    def next_machine_code(self, company_code: str) -> str:
        """Return the next Mnnn code for a company (e.g. 'M003' if M001/M002 exist)."""
        self.invalidate_cache()
        pattern = re.compile(rf"^TF-{re.escape(company_code)}-M(\d+)$", re.IGNORECASE)
        max_n = 0
        for machine_id in self.load():
            m = pattern.match(machine_id)
            if m:
                max_n = max(max_n, int(m.group(1)))
        return f"M{max_n + 1:03d}"

    def get_company_machines(self, company_code: str) -> List[dict]:
        with self._lock:
            wb = openpyxl.load_workbook(self._path, data_only=True)
            ws = wb["Machines"]
            machines = []
            for row_cells in ws.iter_rows(min_row=2, values_only=False):
                company = row_cells[MACHINES_HEADER.index("company_code")].value
                if company == company_code:
                    # Pad short rows (workbooks pre-dating new columns) with None
                    row_vals = [c.value for c in row_cells]
                    while len(row_vals) < len(MACHINES_HEADER):
                        row_vals.append(None)
                    machines.append(
                        {col: row_vals[i] for i, col in enumerate(MACHINES_HEADER)}
                    )
            return machines

    def update_machine(self, machine_id: str, fields: dict) -> bool:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            if "Machines" not in wb.sheetnames:
                return False
            ws = wb["Machines"]
            existing_header = [c.value for c in ws[1]]
            target = machine_id.upper()
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=1).value).strip().upper() == target:
                    for key, val in fields.items():
                        if key in existing_header:
                            ws.cell(row=r, column=existing_header.index(key) + 1, value=val)
                    wb.save(self._path)
                    self.invalidate_cache()
                    return True
        return False
