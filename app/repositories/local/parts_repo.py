"""Local (openpyxl / Excel) implementation of PartsRepository."""

import threading
from typing import List, Optional

import openpyxl

from app.repositories.base import (
    CONSUMABLES_HEADER,
    SPARE_PARTS_HEADER,
    PartsRepository,
    new_item_id,
)

_SHEETS = {
    "spare_parts": ("SpareParts", SPARE_PARTS_HEADER),
    "consumables": ("Consumables", CONSUMABLES_HEADER),
}


class LocalPartsRepository(PartsRepository):
    """Reads/writes spare parts and consumables in the local workbook."""

    def __init__(self, xlsx_path: str):
        self._path = xlsx_path
        self._lock = threading.Lock()

    def next_item_id(self, kind: str) -> str:
        return new_item_id(kind)

    def _ensure_sheet(self, wb, kind: str):
        sheet_name, header = _SHEETS[kind]
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(header)
        return wb[sheet_name]

    def list_items(
        self, kind: str, company_code: str, machine_id: Optional[str] = None
    ) -> List[dict]:
        sheet_name, header = _SHEETS[kind]
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            record = dict(zip(header, row))
            if record.get("company_code") != company_code:
                continue
            if machine_id is not None and record.get("machine_id") != machine_id:
                continue
            results.append(record)
        return results

    def get_item(self, kind: str, item_id: str) -> Optional[dict]:
        sheet_name, header = _SHEETS[kind]
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == item_id:
                return dict(zip(header, row))
        return None

    def add_item(self, kind: str, row: dict) -> None:
        _, header = _SHEETS[kind]
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            ws = self._ensure_sheet(wb, kind)
            ws.append([row.get(col, "") for col in header])
            wb.save(self._path)

    def update_item(self, kind: str, item_id: str, updates: dict) -> bool:
        sheet_name, header = _SHEETS[kind]
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            for row_cells in ws.iter_rows(min_row=2):
                if row_cells[0].value == item_id:
                    for field, value in updates.items():
                        if field in header:
                            row_cells[header.index(field)].value = value
                    wb.save(self._path)
                    return True
        return False

    def delete_item(self, kind: str, item_id: str) -> bool:
        sheet_name, _ = _SHEETS[kind]
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            for row_cells in ws.iter_rows(min_row=2):
                if row_cells[0].value == item_id:
                    ws.delete_rows(row_cells[0].row)
                    wb.save(self._path)
                    return True
        return False
