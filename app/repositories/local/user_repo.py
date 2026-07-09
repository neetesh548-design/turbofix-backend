"""Local (openpyxl / Excel) implementation of UserRepository."""

import threading
from typing import List, Optional

import openpyxl

from app.repositories.base import (
    COMPANIES_HEADER,
    USERS_HEADER,
    UserRepository,
    new_user_id,
)


def _normalize(value) -> str:
    return (str(value) if value is not None else "").strip().lower()


class LocalUserRepository(UserRepository):
    """Reads/writes Users and Companies tabs in the local tracker workbook."""

    def __init__(self, xlsx_path: str):
        self._path = xlsx_path
        self._lock = threading.Lock()

    # ------------------------------------------------------------------
    # ID generation
    # ------------------------------------------------------------------

    def next_user_id(self, company_code: str) -> str:
        return new_user_id(company_code)

    # ------------------------------------------------------------------
    # User CRUD
    # ------------------------------------------------------------------

    def get_by_identifier(self, identifier: str) -> Optional[dict]:
        """Look up a user by phone or email (case-insensitive, whitespace-trimmed)."""
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if "Users" not in wb.sheetnames:
            return None
        ws = wb["Users"]
        target = _normalize(identifier)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            record = dict(zip(USERS_HEADER, row))
            if _normalize(record.get("phone")) == target or _normalize(record.get("email")) == target:
                return record
        return None

    def get_by_id(self, user_id: str) -> Optional[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if "Users" not in wb.sheetnames:
            return None
        ws = wb["Users"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == user_id:
                return dict(zip(USERS_HEADER, row))
        return None

    def add(self, row: dict) -> None:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            if "Users" not in wb.sheetnames:
                ws = wb.create_sheet("Users")
                ws.append(USERS_HEADER)
            else:
                ws = wb["Users"]
            ws.append([row.get(col, "") for col in USERS_HEADER])
            wb.save(self._path)

    def update_password(self, user_id: str, new_password_hash: str) -> bool:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            if "Users" not in wb.sheetnames:
                return False
            ws = wb["Users"]
            hash_col = USERS_HEADER.index("password_hash") + 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == user_id:
                    ws.cell(row=r, column=hash_col, value=new_password_hash)
                    wb.save(self._path)
                    return True
        return False

    # ------------------------------------------------------------------
    # Company CRUD
    # ------------------------------------------------------------------

    def get_company(self, company_code: str) -> Optional[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if "Companies" not in wb.sheetnames:
            return None
        ws = wb["Companies"]
        target = _normalize(company_code)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            record = dict(zip(COMPANIES_HEADER, row))
            if _normalize(record.get("company_code")) == target:
                return record
        return None

    def list_companies(self) -> List[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if "Companies" not in wb.sheetnames:
            return []
        ws = wb["Companies"]
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                out.append(dict(zip(COMPANIES_HEADER, row)))
        return out

    def update_company(self, company_code: str, fields: dict) -> bool:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            if "Companies" not in wb.sheetnames:
                return False
            ws = wb["Companies"]
            existing_header = [c.value for c in ws[1]]
            for col in COMPANIES_HEADER:
                if col not in existing_header:
                    ws.cell(row=1, column=len(existing_header) + 1, value=col)
                    existing_header.append(col)
            target = _normalize(company_code)
            for r in range(2, ws.max_row + 1):
                if _normalize(ws.cell(row=r, column=1).value) == target:
                    for key, value in fields.items():
                        if key in existing_header:
                            ws.cell(row=r, column=existing_header.index(key) + 1, value=value)
                    wb.save(self._path)
                    return True
        return False
