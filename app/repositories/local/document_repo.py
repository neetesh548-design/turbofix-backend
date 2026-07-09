"""Local (openpyxl / Excel) implementation of DocumentRepository."""

import threading
from typing import List, Optional

import openpyxl

from app.repositories.base import (
    DOCUMENTS_HEADER,
    DocumentRepository,
    new_document_id,
)


class LocalDocumentRepository(DocumentRepository):
    """Reads/writes document metadata in the Documents tab of the local workbook."""

    def __init__(self, xlsx_path: str):
        self._path = xlsx_path
        self._lock = threading.Lock()

    def next_document_id(self) -> str:
        return new_document_id()

    def _ensure_sheet(self, wb):
        if "Documents" not in wb.sheetnames:
            ws = wb.create_sheet("Documents")
            ws.append(DOCUMENTS_HEADER)

    def list(self, company_code: str, machine_id: Optional[str] = None) -> List[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if "Documents" not in wb.sheetnames:
            return []
        ws = wb["Documents"]
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            record = dict(zip(DOCUMENTS_HEADER, row))
            if record.get("company_code") != company_code:
                continue
            if machine_id is not None and record.get("machine_id") != machine_id:
                continue
            results.append(record)
        return results

    def get(self, document_id: str) -> Optional[dict]:
        wb = openpyxl.load_workbook(self._path, data_only=True)
        if "Documents" not in wb.sheetnames:
            return None
        ws = wb["Documents"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == document_id:
                return dict(zip(DOCUMENTS_HEADER, row))
        return None

    def add(self, row: dict) -> None:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            self._ensure_sheet(wb)
            ws = wb["Documents"]
            ws.append([row.get(col, "") for col in DOCUMENTS_HEADER])
            wb.save(self._path)

    def delete(self, document_id: str) -> bool:
        with self._lock:
            wb = openpyxl.load_workbook(self._path)
            if "Documents" not in wb.sheetnames:
                return False
            ws = wb["Documents"]
            for row_cells in ws.iter_rows(min_row=2):
                if row_cells[0].value == document_id:
                    ws.delete_rows(row_cells[0].row)
                    wb.save(self._path)
                    return True
        return False
