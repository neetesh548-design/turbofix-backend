"""Reads/writes the Users tab of the tracker workbook. Mirrors store_local.py's
shape (same workbook, same "local xlsx now, pluggable to Sheets later" approach) but
kept as its own module since login lookups are a different access pattern (by phone
or email, not by machine_id) and, unlike Machines, must never be served from a stale
cache - a password change or role change should take effect on the very next login.
"""

import secrets
from datetime import datetime, timezone
from typing import Optional

import openpyxl

from app import config

_USERS_HEADER = ["user_id", "company_code", "name", "phone", "email", "role", "password_hash", "created_at"]
_COMPANIES_HEADER = ["company_code", "company_name", "admin_contact_phone", "onboarded_date"]


def next_user_id(company_code: str) -> str:
    return f"U-{company_code}-{datetime.now(timezone.utc):%Y%m%d%H%M%S}-{secrets.token_hex(2)}"


def _normalize(value: Optional[str]) -> str:
    return (value or "").strip().lower()


def get_user_by_identifier(identifier: str) -> Optional[dict]:
    """Looks a user up by phone or email (case-insensitive, whitespace-trimmed).
    Returns None if the Users tab doesn't exist yet (e.g. an older tracker file) or
    no row matches."""
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if "Users" not in wb.sheetnames:
        return None
    ws = wb["Users"]

    target = _normalize(identifier)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        record = dict(zip(_USERS_HEADER, row))
        if _normalize(record.get("phone")) == target or _normalize(record.get("email")) == target:
            return record
    return None


def get_user_by_id(user_id: str) -> Optional[dict]:
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if "Users" not in wb.sheetnames:
        return None
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == user_id:
            return dict(zip(_USERS_HEADER, row))
    return None


def get_company(company_code: str) -> Optional[dict]:
    """Looks up a row in the Companies tab by company_code (case-insensitive,
    whitespace-trimmed) - used by self-service signup to verify a company's
    admin_contact_phone before creating a supervisor account under it."""
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if "Companies" not in wb.sheetnames:
        return None
    ws = wb["Companies"]
    target = _normalize(company_code)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        record = dict(zip(_COMPANIES_HEADER, row))
        if _normalize(record.get("company_code")) == target:
            return record
    return None


def add_user(row: dict) -> None:
    """row keys: user_id, company_code, name, phone, email, role, password_hash,
    created_at. Used by admin onboarding tooling (scripts/create_user.py) for
    owner/maintenance_head logins, and by POST /auth/signup for self-service
    supervisor (read-only) accounts - the two are the same storage function, the
    role restriction lives in the caller, not here."""
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
    if "Users" not in wb.sheetnames:
        ws = wb.create_sheet("Users")
        ws.append(_USERS_HEADER)
    else:
        ws = wb["Users"]
    ws.append([row.get(col, "") for col in _USERS_HEADER])
    wb.save(config.TRACKER_XLSX_PATH)
