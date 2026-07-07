"""Spare parts (BOM) and consumables lists, one tab each in the tracker workbook.
Both are simple per-machine inventories a maintenance head keeps up to date, so they
share the same shape/behavior - implemented once here and used by two thin routers.
"""

import secrets
import threading
from datetime import datetime, timezone
from typing import List, Optional

import openpyxl

from app import config

_lock = threading.Lock()

_ID_PREFIX = {"spare_parts": "SP", "consumables": "CON"}


def next_item_id(kind: str) -> str:
    prefix = _ID_PREFIX[kind]
    return f"{prefix}-{datetime.now(timezone.utc):%Y%m%d%H%M%S}-{secrets.token_hex(2)}"

SPARE_PARTS_HEADER = [
    "part_id", "company_code", "machine_id", "part_name", "part_number",
    "quantity_on_hand", "unit", "reorder_level", "supplier", "notes",
]
CONSUMABLES_HEADER = [
    "consumable_id", "company_code", "machine_id", "name",
    "quantity_on_hand", "unit", "reorder_level", "notes",
]

_SHEETS = {
    "spare_parts": ("SpareParts", SPARE_PARTS_HEADER),
    "consumables": ("Consumables", CONSUMABLES_HEADER),
}


def _ensure_sheet(wb, kind: str):
    sheet_name, header = _SHEETS[kind]
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(header)
    return wb[sheet_name]


def list_items(kind: str, company_code: str, machine_id: Optional[str] = None) -> List[dict]:
    sheet_name, header = _SHEETS[kind]
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        record = dict(zip(header, row))
        if record["company_code"] != company_code:
            continue
        if machine_id is not None and record["machine_id"] != machine_id:
            continue
        results.append(record)
    return results


def get_item(kind: str, item_id: str) -> Optional[dict]:
    sheet_name, header = _SHEETS[kind]
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == item_id:
            return dict(zip(header, row))
    return None


def add_item(kind: str, row: dict) -> None:
    _, header = _SHEETS[kind]
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        ws = _ensure_sheet(wb, kind)
        ws.append([row.get(col, "") for col in header])
        wb.save(config.TRACKER_XLSX_PATH)


def update_item(kind: str, item_id: str, updates: dict) -> bool:
    """Overwrites only the given fields on the matching row. Returns True if found."""
    sheet_name, header = _SHEETS[kind]
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        for row_cells in ws.iter_rows(min_row=2):
            if row_cells[0].value == item_id:
                for field, value in updates.items():
                    if field in header:
                        row_cells[header.index(field)].value = value
                wb.save(config.TRACKER_XLSX_PATH)
                return True
        return False


def delete_item(kind: str, item_id: str) -> bool:
    sheet_name, _ = _SHEETS[kind]
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        for row_cells in ws.iter_rows(min_row=2):
            if row_cells[0].value == item_id:
                ws.delete_rows(row_cells[0].row)
                wb.save(config.TRACKER_XLSX_PATH)
                return True
        return False
