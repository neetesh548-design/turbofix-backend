"""Reads/writes the Users tab of the tracker workbook. Mirrors store_local.py's
shape (same workbook, same "local xlsx now, pluggable to Sheets later" approach) but
kept as its own module since login lookups are a different access pattern (by phone
or email, not by machine_id) and, unlike Machines, must never be served from a stale
cache - a password change or role change should take effect on the very next login.
"""

import secrets
from datetime import datetime, timezone
from typing import Optional

import openpyxl

from app import config

_USERS_HEADER = ["user_id", "company_code", "name", "phone", "email", "role", "password_hash", "created_at"]


def next_user_id(company_code: str) -> str:
    return f"U-{company_code}-{datetime.now(timezone.utc):%Y%m%d%H%M%S}-{secrets.token_hex(2)}"


def _normalize(value: Optional[str]) -> str:
    return (value or "").strip().lower()


def get_user_by_identifier(identifier: str) -> Optional[dict]:
    """Looks a user up by phone or email (case-insensitive, whitespace-trimmed).
    Returns None if the Users tab doesn't exist yet (e.g. an older tracker file) or
    no row matches."""
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if "Users" not in wb.sheetnames:
        return None
    ws = wb["Users"]

    target = _normalize(identifier)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        record = dict(zip(_USERS_HEADER, row))
        if _normalize(record.get("phone")) == target or _normalize(record.get("email")) == target:
            return record
    return None


def get_user_by_id(user_id: str) -> Optional[dict]:
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if "Users" not in wb.sheetnames:
        return None
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == user_id:
            return dict(zip(_USERS_HEADER, row))
    return None


def add_user(row: dict) -> None:
    """row keys: user_id, company_code, name, phone, email, role, password_hash,
    created_at. Used by onboarding tooling, not exposed over the API - creating a
    login is a deliberate admin action, not a self-serve signup."""
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
    if "Users" not in wb.sheetnames:
        ws = wb.create_sheet("Users")
        ws.append(_USERS_HEADER)
    else:
        ws = wb["Users"]
    ws.append([row.get(col, "") for col in _USERS_HEADER])
    wb.save(config.TRACKER_XLSX_PATH)
