"""Metadata for uploaded machine documents (manuals, circuit diagrams, hydraulic
diagrams, etc.) - one row per document in the tracker's Documents tab. The actual
file bytes live in app.file_storage; this module only tracks what exists and where.
"""

import secrets
import threading
from datetime import datetime, timezone
from typing import List, Optional

import openpyxl

from app import config

_lock = threading.Lock()


def next_document_id() -> str:
    return f"DOC-{datetime.now(timezone.utc):%Y%m%d%H%M%S}-{secrets.token_hex(2)}"

_DOCUMENTS_HEADER = [
    "document_id", "company_code", "machine_id", "category", "title",
    "file_name", "storage_path", "uploaded_by", "uploaded_at",
]

# Free-text machine documentation categories. Kept open-ended (not a strict enum in
# the workbook) since factories will inevitably have documents that don't fit neatly.
DOCUMENT_CATEGORIES = ["manual", "circuit_diagram", "hydraulic_diagram", "spare_parts_catalog", "other"]


def _ensure_sheet(wb) -> None:
    if "Documents" not in wb.sheetnames:
        ws = wb.create_sheet("Documents")
        ws.append(_DOCUMENTS_HEADER)


def list_documents(company_code: str, machine_id: Optional[str] = None) -> List[dict]:
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if "Documents" not in wb.sheetnames:
        return []
    ws = wb["Documents"]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        record = dict(zip(_DOCUMENTS_HEADER, row))
        if record["company_code"] != company_code:
            continue
        if machine_id is not None and record["machine_id"] != machine_id:
            continue
        results.append(record)
    return results


def get_document(document_id: str) -> Optional[dict]:
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=True)
    if "Documents" not in wb.sheetnames:
        return None
    ws = wb["Documents"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == document_id:
            return dict(zip(_DOCUMENTS_HEADER, row))
    return None


def add_document(row: dict) -> None:
    """row keys match _DOCUMENTS_HEADER above."""
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        _ensure_sheet(wb)
        ws = wb["Documents"]
        ws.append([row.get(col, "") for col in _DOCUMENTS_HEADER])
        wb.save(config.TRACKER_XLSX_PATH)


def delete_document(document_id: str) -> bool:
    """Removes the matching row. Returns True if a row was found and removed."""
    with _lock:
        wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH)
        if "Documents" not in wb.sheetnames:
            return False
        ws = wb["Documents"]
        for row_cells in ws.iter_rows(min_row=2):
            if row_cells[0].value == document_id:
                ws.delete_rows(row_cells[0].row)
                wb.save(config.TRACKER_XLSX_PATH)
                return True
        return False
