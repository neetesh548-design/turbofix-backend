"""Creates the live Google Sheet ("TurboFix-Tracker-Live") from TurboFix-Tracker.xlsx.

One-time Phase 3 setup: copies every tab (values + formulas) into a new spreadsheet
owned by the service account, then shares it with the given email as editor so a
human can watch tickets arrive. Re-running reuses an existing sheet with the same
title instead of creating duplicates.

Usage:
    python -m scripts.create_live_sheet <share_email>

Requires GOOGLE_SERVICE_ACCOUNT_FILE in .env (TRACKER_XLSX_PATH optional, defaults
to ../TurboFix-Tracker.xlsx). Prints the sheet ID to put in GOOGLE_SHEET_ID.
"""

import sys

import gspread
import openpyxl
from google.oauth2.service_account import Credentials

from app import config

SHEET_TITLE = "TurboFix-Tracker-Live"
_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def _cell_value(cell):
    if cell.value is None:
        return ""
    return str(cell.value) if not isinstance(cell.value, (int, float, str)) else cell.value


def main() -> None:
    if len(sys.argv) != 2:
        sys.exit("usage: python -m scripts.create_live_sheet <share_email>")
    share_email = sys.argv[1]

    creds = Credentials.from_service_account_file(
        config.GOOGLE_SERVICE_ACCOUNT_FILE, scopes=_SCOPES
    )
    client = gspread.authorize(creds)

    try:
        spreadsheet = client.open(SHEET_TITLE)
        print(f"reusing existing spreadsheet '{SHEET_TITLE}'")
    except gspread.SpreadsheetNotFound:
        spreadsheet = client.create(SHEET_TITLE)
        print(f"created spreadsheet '{SHEET_TITLE}'")

    # data_only=False keeps formulas (e.g. the Dashboard tab) as "=..." strings,
    # which USER_ENTERED below re-interprets as live formulas in Google Sheets.
    wb = openpyxl.load_workbook(config.TRACKER_XLSX_PATH, data_only=False)
    existing_titles = {ws.title for ws in spreadsheet.worksheets()}

    for sheet_name in wb.sheetnames:
        src = wb[sheet_name]
        rows = [[_cell_value(c) for c in row] for row in src.iter_rows()]
        if not rows:
            rows = [[""]]

        if sheet_name in existing_titles:
            ws = spreadsheet.worksheet(sheet_name)
            ws.clear()
        else:
            ws = spreadsheet.add_worksheet(
                title=sheet_name, rows=max(len(rows) + 200, 400), cols=len(rows[0]) + 5
            )
        # RAW keeps phone numbers and IDs as text (USER_ENTERED would coerce them to
        # numbers, which breaks string comparisons in the sheets stores). Formulas
        # are re-applied in a second pass below, since RAW would leave them as text.
        ws.update(rows, value_input_option="RAW")

        formula_cells = [
            {"range": gspread.utils.rowcol_to_a1(r + 1, c + 1), "values": [[val]]}
            for r, row_vals in enumerate(rows)
            for c, val in enumerate(row_vals)
            if isinstance(val, str) and val.startswith("=")
        ]
        if formula_cells:
            ws.batch_update(formula_cells, value_input_option="USER_ENTERED")
        print(f"  wrote tab '{sheet_name}' ({len(rows)} rows, {len(formula_cells)} formulas)")

    # drop gspread's default empty tab if it's still around
    for ws in spreadsheet.worksheets():
        if ws.title == "Sheet1" and ws.title not in wb.sheetnames:
            spreadsheet.del_worksheet(ws)

    spreadsheet.share(share_email, perm_type="user", role="writer", notify=False)
    print(f"shared with {share_email} (editor)")

    print()
    print(f"GOOGLE_SHEET_ID={spreadsheet.id}")
    print(f"URL: https://docs.google.com/spreadsheets/d/{spreadsheet.id}")


if __name__ == "__main__":
    main()
